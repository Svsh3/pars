from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


def save_to_excel(rows: list[dict], filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "WB Results"

    # --- Стили ---
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="5C2D91")  # фиолетовый WB
    query_font = Font(name="Arial", bold=True, color="1A1A1A", size=10)
    query_fill = PatternFill("solid", fgColor="F0E6FF")
    data_font = Font(name="Arial", size=10)
    price_font = Font(name="Arial", bold=True, color="CB11AB", size=10)  # розовый WB

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Заголовок ---
    headers = ["Запрос / Товар", "Бренд", "Цена (₽)", "Ссылка"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 22

    # --- Группируем товары по запросу ---
    from collections import defaultdict
    grouped = defaultdict(list)
    for row in rows:
        grouped[row.get("query", "")].append(row)

    current_row = 2

    for query, products in grouped.items():
        # Строка-разделитель с запросом
        query_cell = ws.cell(row=current_row, column=1, value=f"🔍 {query}")
        query_cell.font = query_font
        query_cell.fill = query_fill
        query_cell.alignment = left
        query_cell.border = border

        for col in range(2, 5):
            c = ws.cell(row=current_row, column=col)
            c.fill = query_fill
            c.border = border

        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4
        )
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Товары запроса
        for product in products:
            name = product.get("name", "—")
            brand = product.get("brand", "—")
            price = product.get("price", 0)
            url = product.get("url", "")

            # Название
            c1 = ws.cell(row=current_row, column=1, value=f"  {name}")
            c1.font = data_font
            c1.alignment = left
            c1.border = border

            # Бренд
            c2 = ws.cell(row=current_row, column=2, value=brand)
            c2.font = data_font
            c2.alignment = center
            c2.border = border

            # Цена
            price_display = f"{price:,}".replace(",", " ") + " ₽" if price else "—"
            c3 = ws.cell(row=current_row, column=3, value=price_display)
            c3.font = price_font
            c3.alignment = center
            c3.border = border

            # Ссылка
            c4 = ws.cell(row=current_row, column=4, value=url)
            c4.font = Font(name="Arial", size=10, color="0070C0", underline="single")
            c4.alignment = left
            c4.border = border
            if url:
                c4.hyperlink = url

            ws.row_dimensions[current_row].height = 18
            current_row += 1

    # --- Ширина столбцов ---
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55

    # --- Закрепить шапку ---
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"[excel_builder] Файл сохранён: {filename}")
