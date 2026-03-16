"""
Excel builder — строит сводную таблицу как в образце:
Строки = товары, Колонки = продавцы, Ячейки = цены
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Цветовая палитра ──────────────────────────────────────────────────────────
C_WB_PURPLE   = "CB11AB"   # фирменный пурпурный WB
C_HEADER_BG   = "2D1654"   # тёмно-фиолетовый — шапка
C_HEADER_FG   = "FFFFFF"
C_QUERY_BG    = "F0E6FF"   # светло-лиловый — строка-разделитель запроса
C_QUERY_FG    = "4A0080"
C_ROW_ODD     = "FFFFFF"
C_ROW_EVEN    = "F9F4FF"
C_PRICE       = "1A7A4A"   # зелёный — цена
C_PRICE_BG    = "E8F8EF"   # фон ячейки с ценой
C_SELLER_HEAD = "3D0C6E"   # тёмный фиолетовый — заголовок продавца
C_BORDER      = "D5B8F0"

def _border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def build_pivot_excel(rows: list[dict]) -> bytes:
    """
    rows: list of {query, name, seller, price, url}
    Returns xlsx file as bytes.
    """

    # ── Собираем данные ───────────────────────────────────────────────────────
    # Уникальные продавцы (в порядке появления)
    sellers_ordered = []
    sellers_set = set()
    for r in rows:
        s = r["seller"]
        if s not in sellers_set:
            sellers_ordered.append(s)
            sellers_set.add(s)

    # Группируем по запросам
    queries_order = []
    query_groups: dict[str, list[dict]] = {}
    for r in rows:
        q = r["query"]
        if q not in query_groups:
            query_groups[q] = []
            queries_order.append(q)
        query_groups[q].append(r)

    # ── Создаём книгу ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "WB Цены"
    ws.sheet_view.showGridLines = False

    b = _border()
    b_thick = _border(C_WB_PURPLE, "medium")

    total_cols = 1 + len(sellers_ordered)   # A + продавцы

    # ── Строка 1: мегашапка ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 34
    c = ws.cell(row=1, column=1)
    c.value = f"🛍️  WILDBERRIES — Мониторинг цен  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.font  = _font(bold=True, size=13, color=C_HEADER_FG)
    c.fill  = _fill(C_HEADER_BG)
    c.alignment = _align(h="center")
    c.border = b_thick

    # ── Строка 2: подзаголовок статистики ─────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.row_dimensions[2].height = 18
    c = ws.cell(row=2, column=1)
    c.value = (
        f"Запросов: {len(queries_order)}   •   "
        f"Товаров: {len(rows)}   •   "
        f"Продавцов: {len(sellers_ordered)}"
    )
    c.font  = _font(size=9, color="CCBBEE", italic=True)
    c.fill  = _fill("1E0A3C")
    c.alignment = _align(h="center")

    # ── Строка 3: заголовки колонок ───────────────────────────────────────────
    ws.row_dimensions[3].height = 38
    # Колонка А — "Название товара"
    c = ws.cell(row=3, column=1)
    c.value = "Название товара"
    c.font  = _font(bold=True, size=11, color=C_HEADER_FG)
    c.fill  = _fill(C_WB_PURPLE)
    c.alignment = _align(h="center", wrap=True)
    c.border = b

    # Колонки продавцов
    for ci, seller in enumerate(sellers_ordered, start=2):
        c = ws.cell(row=3, column=ci)
        c.value = seller
        c.font  = _font(bold=True, size=9, color=C_HEADER_FG)
        c.fill  = _fill(C_SELLER_HEAD)
        c.alignment = _align(h="center", wrap=True)
        c.border = b

    # ── Данные ────────────────────────────────────────────────────────────────
    current_row = 4

    for q_idx, query in enumerate(queries_order):
        # --- Строка-разделитель запроса ---
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=total_cols
        )
        ws.row_dimensions[current_row].height = 22
        c = ws.cell(row=current_row, column=1)
        c.value = f"🔍  {query}"
        c.font  = _font(bold=True, size=10, color=C_QUERY_FG)
        c.fill  = _fill(C_QUERY_BG)
        c.alignment = _align(h="left")
        c.border = _border(C_WB_PURPLE, "thin")
        current_row += 1

        # --- Товары этого запроса ---
        group = query_groups[query]
        for item_idx, item in enumerate(group):
            ws.row_dimensions[current_row].height = 30
            row_bg = C_ROW_EVEN if item_idx % 2 == 0 else C_ROW_ODD

            # Столбец А — название товара
            c = ws.cell(row=current_row, column=1)
            c.value = item["name"]
            c.font  = _font(size=9, color="1A0030")
            c.fill  = _fill(row_bg)
            c.alignment = _align(h="left", v="center", wrap=True)
            c.border = b

            # Столбцы продавцов — заполняем None
            for ci, seller in enumerate(sellers_ordered, start=2):
                c = ws.cell(row=current_row, column=ci)
                c.fill   = _fill(row_bg)
                c.border = b
                c.alignment = _align()

            # Заполняем цену в нужной колонке
            seller_col = sellers_ordered.index(item["seller"]) + 2
            c = ws.cell(row=current_row, column=seller_col)
            c.value       = item["price"]
            c.font        = _font(bold=True, size=10, color=C_PRICE)
            c.fill        = _fill(C_PRICE_BG)
            c.number_format = '#,##0 "₽"'
            c.alignment   = _align()
            c.border      = b

            current_row += 1

        # Пустая строка-разделитель между запросами (кроме последнего)
        if q_idx < len(queries_order) - 1:
            ws.row_dimensions[current_row].height = 6
            for ci in range(1, total_cols + 1):
                c = ws.cell(row=current_row, column=ci)
                c.fill = _fill("EDE0FF")
            current_row += 1

    # ── Ширина колонок ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52   # Название товара

    # Ширина колонок продавцов
    max_seller_len = max((len(s) for s in sellers_ordered), default=10)
    col_w = max(12, min(max_seller_len + 2, 22))

    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    # ── Заморозка заголовков ──────────────────────────────────────────────────
    ws.freeze_panes = "B4"

    # ── Второй лист: исходные данные (flat) ───────────────────────────────────
    ws2 = wb.create_sheet("Исходные данные")
    ws2.sheet_view.showGridLines = False

    flat_headers = ["Запрос", "Название товара", "Продавец", "Цена (₽)", "Ссылка"]
    ws2.row_dimensions[1].height = 24
    for ci, h in enumerate(flat_headers, 1):
        c = ws2.cell(row=1, column=ci)
        c.value = h
        c.font  = _font(bold=True, size=10, color=C_HEADER_FG)
        c.fill  = _fill(C_HEADER_BG)
        c.alignment = _align(h="center")
        c.border = b

    for ri, item in enumerate(rows, 2):
        ws2.row_dimensions[ri].height = 20
        bg = C_ROW_EVEN if ri % 2 == 0 else C_ROW_ODD
        for ci, val in enumerate([
            item["query"], item["name"], item["seller"],
            item["price"], item.get("url", "")
        ], 1):
            c = ws2.cell(row=ri, column=ci)
            c.value = val
            c.font  = _font(size=9)
            c.fill  = _fill(bg)
            c.border = b
            c.alignment = _align(h="left", v="center")
            if ci == 4:
                c.number_format = '#,##0 "₽"'
                c.alignment = _align()
                c.font = _font(bold=True, size=9, color=C_PRICE)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 60
    ws2.freeze_panes = "A2"

    # ── Сохраняем в байты ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
